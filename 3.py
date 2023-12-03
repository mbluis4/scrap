import os
import re
import requests
import bs4
from openpyxl import load_workbook
from data import vendordata
import time
import datetime


def getPrices():
    wb = load_workbook(filename='test.xlsx')
    for ws in wb:
        vendor = ws.title
        print(vendor)
        colD = ws['D']
        for item in colD[1:]:
            if len(str(item.value)) != '#N/A':
                ws[f'E{item.row}'] = item.value
    now = datetime.datetime.now().strftime("%d-%m-%Y %H_%M")
    wb.save(f'Nuevos_Precios_{str(now)}.xlsx')
    time.sleep(1)


getPrices()
