import re
import requests
import bs4
from os.path import basename
from openpyxl import load_workbook
from data import vendordata
from send_excel import send_excel
import datetime


def getPrices():
    wb = load_workbook(filename='test.xlsx')
    now = datetime.datetime.now().strftime("%d-%m-%Y %H_%M")
    new_col_title = datetime.datetime.now().strftime("%d-%m-%Y")

    for ws in wb:
        vendor = ws.title
        ws['E1'] = str(new_col_title)
        colD = ws['D']
        for item in colD[1:]:
            if len(str(item.value)) > 10:
                try:
                    response = requests.get(
                        item.value if vendor != 'Todo Griferia' else f'https://{item.value}')
                    print(f'descargando desde {item.value}')
                    response.raise_for_status()
                except requests.exceptions.HTTPError as err:
                    print(err)
                    continue
                print('parsing html')
                s = bs4.BeautifulSoup(response.text, 'html.parser')

                ws[f'E{item.row}'] = parse_page(s, vendor)
    now = datetime.datetime.now().strftime("%d-%m-%Y %H_%M")
    # send file by email
    excel_file = f'Nuevos_Precios_{str(now)}.xlsx'
    wb.save(excel_file)
    return basename(excel_file)
    # time.sleep(1)


def parse_page(s, vendor):
    print('entering parse_page function')
    tags = vendordata[vendor]

    prod_price = s.find(
        tags['price_tag'][0], class_=tags['price_tag'][1])

    price_regex = re.compile(r'\b\d[\d,.]*\b')
    if prod_price is None:
        prod_price_f = 'sin precio'
    elif re.search(price_regex, prod_price.text.strip()) != None:
        prod_price_f = price_regex.search(
            prod_price.text.strip()).group()
    else:
        prod_price_f = 'sin precio'

    return prod_price_f
