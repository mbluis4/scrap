import requests
import os
import time
import re
import bs4
import lxml
from datetime import datetime
from openpyxl import Workbook, load_workbook
from data import vendordatacyber, lines, brands


def lambda_handler(event, context):
    return


def get_page(tienda):
    prod_data = [
        [tienda, 'Marca', 'Linea', 'Nombre', 'Precio', 'Link', 'Cuotas'],
    ]

    def get_urls(tienda):
        base_url = vendordatacyber[tienda]['base_url']
        urls = []
        match tienda:
            case 'Todo Griferia':
                urls.append(base_url)  # 500

            # case 'Sanitarios Arrieta':
            #    for web_page in range(0, 36):  # 0-36
            #        urls.append(f'{base_url}{brand}_Desde_{web_page*50+1}')

            case 'Tucson':
                for web_page in range(1, 11):  # 11
                    urls.append(f'{base_url}{web_page}')

            case 'Blaisten':
                for web_page in range(1, 16):  # 1-16
                    urls.append(f'{base_url}{web_page}')

            # case 'Banchero':
            #    for web_page in range(0, 13):  # 0-13
            #        urls.append(f'{base_url}{brand}_Desde_{web_page*50+1}')
            case 'Bercomat':
                urls.append(base_url)  # 600
             # case 'Acon':
            #    for web_page in range(1, 3):
            #        urls.append(
            #           f'{base_url}{web_page}/?s={brand}&post_type=product&dgwt_wcas=1')

        return urls
    for page in get_urls(tienda):
        try:
            response = requests.get(page)
            print(f'descargando desde {page}')
            response.raise_for_status()
        except requests.exceptions.HTTPError as err:
            print(err)
            break
        print('parsing html')
        s = bs4.BeautifulSoup(response.text, 'lxml')
        if s.find_all('div', class_=vendordatacyber[tienda]['main_tag']) == []:
            print(f'end of {tienda} pages')
            break
        prod_data += parse_page(s, tienda)
        print('next page download in 2 seconds...')
        time.sleep(1)
    return prod_data


def parse_page(s, tienda):
    print('entering parse_page function')
    tags = vendordatacyber[tienda]
    page_data = []
    for item in s.find_all(
            'div', class_=tags['main_tag']):
        prod_link = item.find('a', href=True)
        prod_name = item.find(tags['name_tag'][0], class_=tags['name_tag'][1])
        prod_price = item.find(
            tags['price_tag'][0], class_=tags['price_tag'][1])
        prod_cuotas = item.find('span', class_=tags['cuotas_tag'])
        # price validation
        # print(prod_price)
        if prod_price is None:
            prod_price_f = 'sin precio'
        else:
            price_regex = re.compile(r'\b\d[\d,.]*\b')
            prod_price_f = price_regex.search(prod_price.text.strip()).group()

        def brand_type(name):
            for n in brands:
                if n.lower() in name.lower():
                    return n
            return 'sin marca'

        def line_type(name):
            for n in lines[brand_type(prod_name.text.strip())]:
                if n.lower() in name.lower():
                    return n

        def get_cuotas(prod):
            if prod is None:
                return 0
            else:
                return prod.text.strip()

        match tienda:
            case 'Todo Griferia':
                prod_link_f = f'www.todogriferia.com{prod_link["href"]}'
            case 'Bercomat':
                prod_link_f = f'https://www.familiabercomat.com{prod_link["href"]}'
            case _:
                prod_link_f = prod_link["href"]

        page_data.append([tienda, brand_type(prod_name.text.strip()), line_type(prod_name.text.strip()), prod_name.text.strip(),
                          prod_price_f, prod_link_f, get_cuotas(prod_cuotas)])
    return page_data

# saving to excel file


def save_xls(tienda):
    now = datetime.today().strftime('%d-%m-%Y')

    data = get_page(tienda)
    if f'Precios_{str(now)}.xlsx' in os.listdir('.'):
        wb = load_workbook(filename=f'Precios_{str(now)}.xlsx')
        ws = wb.active
        for row in data[1:]:
            ws.append(row)
        print('saving to file...')
        wb.save(f'Precios_{str(now)}.xlsx')
    else:
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        print('creating and saving to file...')
        wb.save(f'Precios_{str(now)}.xlsx')


# if __name__ == '__main__':
#    for tienda in vendordatacyber.keys():
#        print(f'Descargando precios de {tienda}')
#        save_xls(tienda)

save_xls('Tucson')
