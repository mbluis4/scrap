import requests, time
import bs4
import lxml
from openpyxl import load_workbook


blaisten_alberdi_1 = 'https://www.blaisten.com.ar/buscapagina?PS=50&sl=19ccd66b-b568-43cb-a106-b52f9796f5cd&cc=25&sm=0&ft=alberdi&PageNumber=1'



blaisten_alberdi_total = [
    blaisten_alberdi_1,

]



prod_data = [
    ['Nombre', 'Precio', 'Link']
]

for page in blaisten_alberdi_total:
    try:
        print(f'descargando desde {page}')
        response = requests.get(page)
    except:
        print('Page not found')
        continue

    s = bs4.BeautifulSoup(response.text, 'lxml')
    print('parsing html')
    for item in s.find_all(
            'div', class_='product-card'):
        prod_link = item.find('a', href=True)
        prod_name = item.find_all('div', class_='product-name')
        prod_price = item.find_all(
            'span', class_='final-price')
        

        for prod, price in zip(prod_name, prod_price):
            prod_data.append([prod.text.strip(), price.text.strip(), prod_link['href']])
    print('next page download in 10 s')
    time.sleep(10)

# saving to excel file

wb = load_workbook(filename='blaisten.xlsx')
ws = wb.create_sheet('alberdi')

for row in prod_data:
    ws.append(row)

print('saving to file...')
wb.save('blaisten.xlsx')
