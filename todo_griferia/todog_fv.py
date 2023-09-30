import requests, os, time
import bs4
import lxml
from openpyxl import Workbook, load_workbook
from lines import fv_lines

brands = ['ferrum', 'fv', 'hidromet']

todo_griferia_fv_url = []

prod_data = [
    ['Tienda', 'Marca', 'Linea', 'Nombre', 'Precio', 'Link', ],
]

brand = 'fv'

tienda = 'Todo Griferia'


todo_griferia_fv_url.append(f'https://www.todogriferia.com/on/demandware.store/Sites-TG-Site/default/Search-UpdateGrid?q={brand}&start={1}&sz={12}')


def get_page(url):
    for page in url:
        try:
            print(f'descargando desde {page}')
            response = requests.get(page)
        except:
            print('Page not found')
            continue
    print('parsing html')
    s = bs4.BeautifulSoup(response.text, 'lxml')
    parse_page(s)

def parse_page(s):
    for item in s.find_all(
            'div', class_='product-tile'):
        prod_link = item.find('a', href=True)
        prod_name = item.find('a', class_='link')
        prod_price = item.find('span', class_='sales')
        
        def line_type(name):
            for n in fv_lines:
                if n in name.lower():
                    return n         
        prod_data.append([tienda, brand, line_type(prod_name.text.strip()), prod_name.text.strip(), int(prod_price.text.strip()[1:].replace(',','').replace('.','')), prod_link['href']])
    print('next page download in 5 s')
    #time.sleep(5)

# saving to excel file
def save_xls(data):
    get_page(todo_griferia_fv_url)

    if 'todo_griferia.xlsx' in os.listdir('.'):
        wb = load_workbook(filename='todo_griferia.xlsx')
        ws = wb.active
        for row in data[1:]:
            ws.append(row)
        print('loading file...')
        print('saving to file...')
        wb.save('todo_griferia.xlsx')
            
    else:        
        wb = Workbook()
        ws = wb.active
        ws.title = 'todo_griferia'
        for row in data:
            ws.append(row)
        print('creating...')
        print('saving to file...')
        wb.save('todo_griferia.xlsx')
  
    
save_xls(prod_data)
