import requests, time
import bs4
import lxml
from openpyxl import Workbook, load_workbook
from lines import peirano_lines

tienda ='Foschia'
brand = 'peirano'
foschia_peirano_total = []
prod_data = [
]


for page_number in range(1,14):
    foschia_peirano_total.append(f'https://foschia.com.ar/search/all/peirano?page={page_number}')


for page in foschia_peirano_total:
    try:
        print(f'descargando desde {page}')
        response = requests.get(page)
    except:
        print('Page not found')
        continue

    s = bs4.BeautifulSoup(response.text, 'lxml')
    print('parsing html')
    for item in s.find_all(
            'div', class_='card'):
        prod_link = item.find('a', href=True)
        prod_name = item.find_all('h3', class_='title')
        prod_price = item.find_all(
            'div', class_='price')
        
        def line_type(name):
            for n in peirano_lines:
                if n.lower() in name.lower():
                    return n 
              
        for prod, price in zip(prod_name, prod_price):
            price_1 = price.text.strip()[2:]
            prod_data.append([tienda, brand, line_type(prod.text.strip()), prod.text.strip(), price_1, prod_link['href']])
    print('next page download in 5 seconds')
    time.sleep(5)

# saving to excel file

#wb = Workbook()
#ws = wb.active
#ws.title = 'peirano'

wb = load_workbook(filename='foschia.xlsx')
ws = wb.create_sheet('peirano')

for row in prod_data:
    ws.append(row)

print('saving to file...')
wb.save('foschia.xlsx')
print('TASK COMPLETED!')
