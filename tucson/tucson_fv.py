import requests, time
import bs4
import lxml
from openpyxl import Workbook, load_workbook


tucson_fv_total = []
prod_data = [
    ['Linea', 'Nombre', 'Precio', 'Link',]

]

lineas = [
'california',
'obera',
'kansas',
'alabama',
'temple',
'newport',
'arizona',
'libby',
'dominic',
'puelo',
'urbano',
'alesia',
'vermont',
'denisse',
'epuyen',
'alerce',
'pampa',
'chalten',
'triades',
]

for page_number in range(1,33):
    tucson_fv_total.append(f'https://tienda.tucsonsa.com/search/page/{page_number}/?q=fv&results_only=true&limit=12&theme=amazonas')


for page in tucson_fv_total:
    try:
        print(f'descargando desde {page}')
        response = requests.get(page)
    except:
        print('Page not found')
        continue

    s = bs4.BeautifulSoup(response.text, 'lxml')
    print('parsing html')
    for item in s.find_all(
            'div', class_='item-description'):
        prod_link = item.find('a', href=True)
        prod_name = item.find_all('div', class_='js-item-name')
        prod_price = item.find_all(
            'span', class_='js-price-display')
        
        def line_type(name):
            for n in lineas:
                if n in name.lower():
                    return n 
                    
        for prod, price in zip(prod_name, prod_price):
            prod_data.append([line_type(prod.text.strip()), prod.text.strip(), price.text.strip(), prod_link['href']])
    print('next page download in 10 s')
    time.sleep(10)

# saving to excel file

#wb = Workbook()
#ws = wb.active
#ws.title = 'fv'

wb = load_workbook(filename='tucson.xlsx')
ws = wb.create_sheet('fv')

for row in prod_data:
    ws.append(row)

print('saving to file...')
wb.save('tucson.xlsx')
