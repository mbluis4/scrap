import requests, os, time
import bs4
import lxml
from openpyxl import Workbook, load_workbook
from data.lines import lines

tienda = 'Todo Griferia'

def get_page(brand):
    prod_data = [
    ['Tienda', 'Marca', 'Linea', 'Nombre', 'Precio', 'Link', ],
    ]
    urls = []
    urls.append(f'https://www.todogriferia.com/on/demandware.store/Sites-TG-Site/default/Search-UpdateGrid?q={brand}&start=0&sz=500')

    for page in urls:
        try:
            print(f'descargando desde {page}')
            response = requests.get(page)
        except requests.exceptions.HTTPError as err:
            print(err)
            break
        print('parsing html')
        s = bs4.BeautifulSoup(response.text, 'lxml')
        if s.find_all('a') == []:
            print(f'end of {brand} pages')
            break
        prod_data += parse_page(s, brand)
        print('next page download in 2 seconds...')
        time.sleep(2)
        return parse_page(s, brand)

def parse_page(s, brand):
    page_data=[]
    for item in s.find_all(
            'div', class_='product-tile'):
        prod_link = item.find('a', href=True)
        prod_name = item.find('a', class_='link')
        prod_price = item.find('span', class_='sales')
        
        def line_type(name):
            for n in lines[brand]:
                if n.lower() in name.lower():
                    return n         
        page_data.append([tienda, brand, line_type(prod_name.text.strip()), prod_name.text.strip(), 
                          prod_price.text.strip()[1:], f'www.todogriferia.com{prod_link["href"]}'])
    return page_data
    

# saving to excel file
def save_xls():

    #brands = ['ferrum', 'fv', 'hidromet', 'peirano', 'vite', 'cerro', 'roca', 'ilva', 'tendenza', 'alberdi', ]
    brands = ['fv']
    for brand in brands:
        data = get_page(brand)

        if 'todo_griferia.xlsx' in os.listdir('.'):
            wb = load_workbook(filename='todo_griferia.xlsx')
            ws = wb.active
            for row in data[1:]:
                ws.append(row)
            print('loading file...')
            print('saving to file...')
            wb.save('todo_griferia.xlsx')
                
        else:        
            wb = Workbook()
            ws = wb.active
            ws.title = 'todo_griferia'
            for row in data:
                ws.append(row)
            print('creating...')
            print('saving to file...')
            wb.save('todo_griferia.xlsx')
    print('next page download in 5 seconds...')
    time.sleep(5)
  
    
save_xls()
