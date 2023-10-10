import requests
import time
import bs4
import lxml
from openpyxl import load_workbook

banchero_fv_1 = 'https://www.tienda.bancherosanitarios.com.ar/fv'

banchero_fv_2 = 'https://www.tienda.bancherosanitarios.com.ar/fv_Desde_51_NoIndex_True'

banchero_fv_3 = 'https://www.tienda.bancherosanitarios.com.ar/fv_Desde_101_NoIndex_True'

banchero_fv_4 = 'https://www.tienda.bancherosanitarios.com.ar/fv_Desde_151_NoIndex_True'

banchero_fv_5 = 'https://www.tienda.bancherosanitarios.com.ar/fv_Desde_201_NoIndex_True'

banchero_fv_6 = 'https://www.tienda.bancherosanitarios.com.ar/fv_Desde_251_NoIndex_True'

banchero_fv_7 = 'https://www.tienda.bancherosanitarios.com.ar/fv_Desde_301_NoIndex_True'

banchero_fv_8 = 'https://www.tienda.bancherosanitarios.com.ar/fv_Desde_351_NoIndex_True'

banchero_fv_9 = 'https://www.tienda.bancherosanitarios.com.ar/fv_Desde_401_NoIndex_True'

banchero_fv_10 = 'https://www.tienda.bancherosanitarios.com.ar/fv_Desde_451_NoIndex_True'

banchero_fv_11 = 'https://www.tienda.bancherosanitarios.com.ar/fv_Desde_501_NoIndex_True'

banchero_fv_12 = 'https://www.tienda.bancherosanitarios.com.ar/fv_Desde_551_NoIndex_True'

banchero_fv_total = [
    banchero_fv_1,
    banchero_fv_2,
    banchero_fv_3,
    banchero_fv_4,
    banchero_fv_5,
    banchero_fv_6,
    banchero_fv_7,
    banchero_fv_8,
    banchero_fv_9,
    banchero_fv_10,
    banchero_fv_11,
    banchero_fv_12,
]

banchero_prices = [
    ['Descripción', 'Precio', 'Link']
]

for page in banchero_fv_total:
    try:
        print(f'descargando desde {page}')
        response = requests.get(page)
    except:
        print('Page not found')
        continue

    s = bs4.BeautifulSoup(response.text, 'lxml')
    print('parsing html')
    # finding div container for each product
    for item in s.find_all(
            'div', class_='ui-search-result__content-wrapper shops__result-content-wrapper'):
        prod_link = item.find_all('a', href=True)
        prod_name = item.find_all('h2', class_='ui-search-item__title')
        prod_price = item.find_all(
            'span', class_='andes-money-amount__fraction')
        for prod, price, link in zip(prod_name, prod_price, prod_link):
            banchero_prices.append([prod.text, price.text, link['href']])
    print('10 seconds for the next page to parse')
time.sleep(10)


wb = load_workbook(filename='banchero.xlsx')
ws1 = wb.create_sheet('fv')

for row in banchero_prices:
    ws1.append(row)

print(f'saving to file')
wb.save('banchero.xlsx')
