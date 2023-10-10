import requests
import time
import bs4
import lxml
from openpyxl import Workbook, load_workbook


banchero_ferrum_1 = 'https://www.tienda.bancherosanitarios.com.ar/ferrum'

banchero_ferrum_2 = 'https://www.tienda.bancherosanitarios.com.ar/ferrum_Desde_51_NoIndex_True'

banchero_ferrum_3 = 'https://www.tienda.bancherosanitarios.com.ar/ferrum_Desde_101_NoIndex_True'

banchero_ferrum_4 = 'https://www.tienda.bancherosanitarios.com.ar/ferrum_Desde_151_NoIndex_True'

banchero_ferrum_5 = 'https://www.tienda.bancherosanitarios.com.ar/ferrum_Desde_201_NoIndex_True'

banchero_ferrum_6 = 'https://www.tienda.bancherosanitarios.com.ar/ferrum_Desde_251_NoIndex_True'

banchero_ferrum_7 = 'https://www.tienda.bancherosanitarios.com.ar/ferrum_Desde_301_NoIndex_True'

banchero_ferrum_total = [
    banchero_ferrum_1,
    banchero_ferrum_2,
    banchero_ferrum_3,
    banchero_ferrum_4,
    banchero_ferrum_5,
    banchero_ferrum_6,
    banchero_ferrum_7,
]

banchero_prices = [
    ['Nombre', 'Precio', 'Link']
]

for page in banchero_ferrum_total:
    try:
        print(f'descargando desde {page}')
        response = requests.get(page)
    except:
        print('Page not found')
        continue

    s = bs4.BeautifulSoup(response.text, 'lxml')
    print('parsing html')
    for item in s.find_all(
            'div', class_='ui-search-result__content-wrapper shops__result-content-wrapper'):
        prod_link = item.find_all('a', href=True)
        prod_name = item.find_all('h2', class_='ui-search-item__title')
        prod_price = item.find_all(
            'span', class_='andes-money-amount__fraction')
        for prod, price, link in zip(prod_name, prod_price, prod_link):
            banchero_prices.append([prod.text, price.text, link['href']])
    print('waiting 10s for the next page to download..')
    time.sleep(10)

# saving to excel file

wb = load_workbook(filename='banchero.xlsx')
ws1 = wb.create_sheet('banchero')

for row in banchero_prices:
    ws1.append(row)

print('saving to file...')
wb.save('banchero.xlsx')
