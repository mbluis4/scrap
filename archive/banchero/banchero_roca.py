import requests
import time
import bs4
from openpyxl import load_workbook

banchero_roca_1 = 'https://www.tienda.bancherosanitarios.com.ar/roca'

banchero_roca_2 = 'https://www.tienda.bancherosanitarios.com.ar/roca_Desde_51_NoIndex_True'

banchero_roca_3 = 'https://www.tienda.bancherosanitarios.com.ar/roca_Desde_101_NoIndex_True'

banchero_roca_4 = 'https://www.tienda.bancherosanitarios.com.ar/roca_Desde_151_NoIndex_True'

banchero_roca_5 = 'https://www.tienda.bancherosanitarios.com.ar/roca_Desde_201_NoIndex_True'

banchero_roca_6 = 'https://www.tienda.bancherosanitarios.com.ar/roca_Desde_251_NoIndex_True'

banchero_roca_7 = 'https://www.tienda.bancherosanitarios.com.ar/roca_Desde_301_NoIndex_True'

banchero_roca_8 = 'https://www.tienda.bancherosanitarios.com.ar/roca_Desde_351_NoIndex_True'

banchero_roca_9 = 'https://www.tienda.bancherosanitarios.com.ar/roca_Desde_401_NoIndex_True'

banchero_roca_10 = 'https://www.tienda.bancherosanitarios.com.ar/roca_Desde_451_NoIndex_True'

banchero_roca_11 = 'https://www.tienda.bancherosanitarios.com.ar/roca_Desde_501_NoIndex_True'

banchero_roca_total = [
    banchero_roca_1,
    banchero_roca_2,
    banchero_roca_3,
    banchero_roca_4,
    banchero_roca_5,
    banchero_roca_6,
    banchero_roca_7,
    banchero_roca_8,
    banchero_roca_9,
    banchero_roca_10,
    banchero_roca_11,
]

banchero_prices = [
    ['Descripción', 'Precio', 'Link']
]

for page in banchero_roca_total:
    try:
        print(f'descargando desde {page}')
        response = requests.get(page)
    except:
        print('Page not found')
        continue

    s = bs4.BeautifulSoup(response.text, 'lxml')
    print('parsing html')
    # finding div container for each product
    for item in s.find_all(
            'div', class_='ui-search-result__content-wrapper shops__result-content-wrapper'):
        prod_link = item.find_all('a', href=True)
        prod_name = item.find_all('h2', class_='ui-search-item__title')
        prod_price = item.find_all(
            'span', class_='andes-money-amount__fraction')
        for prod, price, link in zip(prod_name, prod_price, prod_link):
            banchero_prices.append([prod.text, price.text, link['href']])
    print('10 seconds for the next page to parse')
    time.sleep(10)

wb = load_workbook(filename='banchero.xlsx')

ws1 = wb.create_sheet('roca')


for row in banchero_prices:
    ws1.append(row)

wb.save('banchero.xlsx')
