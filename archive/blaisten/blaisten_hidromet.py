import requests, time
import bs4
import lxml
from openpyxl import Workbook, load_workbook


blaisten_hidromet_total = []
prod_data = [
    ['Nombre', 'Precio', 'Link']
]

for page_number in range(1,8):
    blaisten_hidromet_total.append(f'https://www.blaisten.com.ar/buscapagina?fq=B%3a35&PS=25&sl=19ccd66b-b568-43cb-a106-b52f9796f5cd&cc=25&sm=0&PageNumber={page_number}')


for page in blaisten_hidromet_total:
    try:
        print(f'descargando desde {page}')
        response = requests.get(page)
    except:
        print('Page not found')
        continue

    s = bs4.BeautifulSoup(response.text, 'lxml')
    print('parsing html')
    for item in s.find_all(
            'div', class_='product-card'):
        prod_link = item.find('a', href=True)
        prod_name = item.find_all('div', class_='product-name')
        prod_price = item.find_all(
            'span', class_='final-price')

        for prod, price in zip(prod_name, prod_price):
            prod_data.append([prod.text.strip(), price.text.strip(), prod_link['href']])
    print('next page download in 10 s')
    time.sleep(10)

# saving to excel file

#wb = Workbook()
#ws = wb.active
#ws.title = 'ferrum'

wb = load_workbook(filename='blaisten.xlsx')
ws = wb.create_sheet('hidromet')

for row in prod_data:
    ws.append(row)

print('saving to file...')
wb.save('blaisten.xlsx')
