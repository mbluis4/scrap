import requests, time
import bs4
import lxml
from openpyxl import Workbook, load_workbook


tucson_peirano_total = []
prod_data = [
    ['Linea', 'Nombre', 'Precio', 'Link',]

]

lineas = [
'adra',
'toledo',
'malba',
'santander',
'valencia',
'bilbao',
'lorca',
'lugo',
'lago',
'castilla',
'marbella',
'mallorca',
'mora',
'soria',
'black',
'vera',

]

for page_number in range(1,6):
    tucson_peirano_total.append(f'https://tienda.tucsonsa.com/search/page/{page_number}/?q=peirano&results_only=true&limit=12&theme=amazonas')


for page in tucson_peirano_total:
    try:
        print(f'descargando desde {page}')
        response = requests.get(page)
    except:
        print('Page not found')
        continue

    s = bs4.BeautifulSoup(response.text, 'lxml')
    print('parsing html')
    for item in s.find_all(
            'div', class_='item-description'):
        prod_link = item.find('a', href=True)
        prod_name = item.find_all('div', class_='js-item-name')
        prod_price = item.find_all(
            'span', class_='js-price-display')
        
        def line_type(name):
            for n in lineas:
                if n in name.lower():
                    return n 
                    
        for prod, price in zip(prod_name, prod_price):
            prod_data.append([line_type(prod.text.strip()), prod.text.strip(), price.text.strip(), prod_link['href']])
    print('next page download in 10 s')
    time.sleep(10)

# saving to excel file

#wb = Workbook()
#ws = wb.active
#ws.title = 'peirano'

wb = load_workbook(filename='tucson.xlsx')
ws = wb.create_sheet('peirano')

for row in prod_data:
    ws.append(row)

print('saving to file...')
wb.save('tucson.xlsx')
