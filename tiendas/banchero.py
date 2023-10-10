import requests, os, time
import bs4
import lxml
from openpyxl import Workbook, load_workbook
from data.lines import lines

tienda = 'Banchero'

def get_page(brand):
    base_url = 'https://www.tienda.bancherosanitarios.com.ar/'
    urls = []
    prod_data = [
    ['Tienda', 'Marca', 'Linea', 'Nombre', 'Precio', 'Link',],
    ]

    for web_page in range(0,13):
        if web_page == 0:
            urls.append(f'{base_url}{brand}_Desde_1_NoIndex_True')
            continue
        urls.append(f'{base_url}{brand}_Desde_{web_page*50+1}_NoIndex_True')   

    for page in urls:       
        try:
            response = requests.get(page)
            print(f'descargando desde {page}')
            response.raise_for_status()
        except requests.exceptions.HTTPError as err:
            print(err)
            break
        s = bs4.BeautifulSoup(response.text, 'lxml')
        if s.find_all('a') == []:
            print(f'end of {brand} pages')
            break
        print('parsing html')
        prod_data += parse_page(s, brand)
        print('next page download in 2 seconds...')
        time.sleep(2)
    return prod_data
    

def parse_page(s, brand):
    page_data = []   
    for item in s.find_all(
            'div', class_='ui-search-result__content-wrapper shops__result-content-wrapper'):
        prod_link = item.find('a', href=True)
        prod_name = item.find('h2', class_='ui-search-item__title')
        prod_price = item.find(
            'span', class_='andes-money-amount__fraction')
        if prod_price is None:
            prod_price_f = 'sin precio'
        else:
            prod_price_f = prod_price.text.strip() 
        def line_type(name):
            for n in lines[brand]:
                if n.lower() in name.lower():
                    return n
                     
        page_data.append([tienda, brand, line_type(prod_name.text.strip()), prod_name.text.strip(), 
                          prod_price_f, prod_link["href"], ])
    return page_data
    
# saving to excel file
def save_xls():

    brands = ['ferrum', 'fv', 'roca',]
   
    
    for brand in brands:
        data = get_page(brand)

        if f'{tienda}.xlsx' in os.listdir('.'):
            wb = load_workbook(filename=f'{tienda}.xlsx')
            ws = wb.active
            for row in data[1:]:
                ws.append(row)
            print('loading file...')
            print('saving to file...')
            wb.save(f'{tienda}.xlsx')
                
        else:        
            wb = Workbook()
            ws = wb.active
            ws.title = f'{tienda}'
            for row in data:
                ws.append(row)
            print('creating...')
            print('saving to file...')
            wb.save(f'{tienda}.xlsx')
        
  
    
save_xls()
