import requests, os, time, datetime, re
import bs4
import lxml
from openpyxl import Workbook, load_workbook
from data.lines import lines
from data.vendors import vendordata


def get_page(tienda, brand):
    prod_data = [
    [tienda, 'Marca', 'Linea', 'Nombre', 'Precio', 'Link', 'Cuotas' ],
    ]

    def get_urls(tienda, brand):
        base_url = vendordata[tienda]['base_url']
        urls=[]
        if tienda == 'Todo Griferia':
            urls.append(f'{base_url}{brand}&start=0&sz=500')
            return urls
        for web_page in range(0,2):
            urls.append(f'{base_url}{brand}_Desde_{web_page*50+1}')   
        return urls

    for page in get_urls(tienda, brand):       
        try:
            response = requests.get(page)
            print(f'descargando desde {page}')
            response.raise_for_status()
        except requests.exceptions.HTTPError as err:
            print(err)
            break
        print('parsing html')
        s = bs4.BeautifulSoup(response.text, 'lxml')
        if s.find_all('a') == []:
            print(f'end of {brand} pages')
            break
        prod_data += parse_page(s, brand, tienda)
        print('next page download in 2 seconds...')
        time.sleep(2)
    return prod_data
    

def parse_page(s, brand, tienda):
    page_data = []   
    for item in s.find_all(
            'div', class_='ui-search-result__content-wrapper shops__result-content-wrapper'):
        prod_link = item.find('a', href=True)
        prod_name = item.find('h2', class_='ui-search-item__title')
        prod_price = item.find(
            'span', class_='andes-money-amount__fraction')
        prod_cuotas = item.find('span', class_="ui-search-item__group__element shops__items-group-details ui-search-installments ui-search-color--LIGHT_GREEN")
        
        #price validation
        if prod_price is None:
            prod_price_f = 'sin precio'
        else:
            price_regex = re.compile(r'\w+[,|.]\w+')
            prod_price_f = price_regex.search(prod_price.text).group()    

        def line_type(name):
            for n in lines[brand]:

                if n.lower() in name.lower():
                    return n
        def get_cuotas(prod):
            if prod is None:
                return 0
            else:
                return prod.text.strip()
                     
        page_data.append([tienda, brand, line_type(prod_name.text.strip()), prod_name.text.strip(), 
                          prod_price_f, prod_link["href"], get_cuotas(prod_cuotas)])
    return page_data
    
# saving to excel file
def save_xls(tienda):
    today = datetime.date.today()

    #brands = ['ferrum', 'fv', 'hidromet', 'peirano', 'vite', 'cerro', 'roca', 'ilva', 'tendenza', 'alberdi', ]
    brands = ['fv']
    
    for brand in brands:
        data = get_page(tienda, brand)

        if 'Precios.xlsx' in os.listdir('.'):
            wb = load_workbook(filename='Precios.xlsx')
            ws = wb.active
            for row in data[1:]:
                ws.append(row)
            print('loading file...')
            print('saving to file...')
            wb.save('Precios.xlsx')
                
        else:        
            wb = Workbook()
            ws = wb.active
            for row in data:
                ws.append(row)
            print('creating...')
            print('saving to file...')
            wb.save('Precios.xlsx')

tiendas = ['Sanitarios Arrieta',]
for tienda in tiendas:
    save_xls(tienda)
