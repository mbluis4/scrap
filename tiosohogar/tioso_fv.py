import requests
import time
import bs4
import lxml
from openpyxl import Workbook, load_workbook

tioso_fv_url = []

lineas = [
    'california',
    'allegro',
    'arizona',
    'chalten',
    'alesia',
    'dominic',
    'vermont',
    'temple',
    'margot',
    'puelo',
    'malena',
    'chess',
    'libby',
    'denisse',
    'triades',
    'epuyen',
    'alerce',
    'oregon',
]

prod_data = [
    ['Tienda', 'Marca', 'Linea', 'Nombre', 'Precio', 'Link', 'Cuotas']
]
brand = 'fv'

base_url = 'https://www.tiosohogar.com/'

for page in range(0,5):
    if page == 0:
        tioso_fv_url.append(f'{base_url}{brand}_Desde_1_NoIndex_True')
        continue
    tioso_fv_url.append(f'{base_url}{brand}_Desde_{page*50+1}_NoIndex_True')   

for page in tioso_fv_url:
    try:
        print(f'descargando desde {page}')
        response = requests.get(page)
    except:
        print('Page not found')
        continue

    s = bs4.BeautifulSoup(response.text, 'lxml')
    print('parsing html')
    for item in s.find_all(
            'div', class_='ui-search-result__content-wrapper shops__result-content-wrapper'):
        prod_link = item.find_all('a', href=True)
        prod_name = item.find_all('h2', class_='ui-search-item__title')
        prod_price = item.find_all(
            'span', class_='andes-money-amount__fraction')
        prod_cuotas = item.find('span', class_="ui-search-item__group__element shops__items-group-details ui-search-installments ui-search-color--LIGHT_GREEN")

        def get_cuotas(prod):
            if prod is None:
                return 0
            else:
                return prod.text.strip()
            
                
        def line_type(name):
            for n in lineas:
                if n.lower() in name.lower():
                    return n 
        for prod, price, link in zip(prod_name, prod_price, prod_link):
            prod_data.append(['Tiosohogar', brand, line_type(prod.text.strip()),
                               prod.text.strip(), price.text.strip(), link['href'], get_cuotas(prod_cuotas)])
    print('waiting 10s for the next page to download..')
    time.sleep(5)

# saving to excel file

#wb = Workbook()
#ws = wb.active
#ws.title = 'Tiosohogar'

wb = load_workbook(filename='tiosohogar.xlsx')
ws = wb.active

for row in prod_data:
    ws.append(row)

print('saving to file...')
wb.save('tiosohogar.xlsx')
